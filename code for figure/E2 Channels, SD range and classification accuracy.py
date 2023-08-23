import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook


# wb = Workbook()
filename = 'E2 Bushy/bushy'
wb = load_workbook(f'../tab/{filename}.xlsx')
ws = wb['Sheet4']

# rows = ws.max_row
rows = [np.arange(2,12),np.arange(12,22)]
cols = [3,7]

result1=[]
result2=[]
for row_group in rows:
    for ind_col in cols:
        temp1 = []
        temp2 = []
        for ind_row in row_group:
            temp1.append(ws.cell(ind_row, ind_col).value)
            temp2.append(ws.cell(ind_row, ind_col+1).value)
        result1.append(temp1)
        result2.append(temp2)
result1=np.array(result1)
result2=np.array(result2)

fig = plt.figure(figsize=(8,6))

fig.show()
fontsize=8
x=[0.3,0.5,0.8,1,2,3,5,8,10,20]
# title_lis=['(a)','(b)','(c)','(d)','(e)','(f)']
color_list=['r','k','gold','g','royalblue','deeppink','purple']
label_list=['100 channels','200 channels','300 channels','500 channels','800 channels','1000 channels']
linestyle_list = ['solid','dotted','dashed','dashdot',(0, (1, 1)),(0, (5, 10)),(0, (3, 10, 1, 10, 1, 10))]

axe1 = fig.add_subplot(1, 2, 1)
axe2 = fig.add_subplot(1, 2, 2)
for i in range(0,4):

    axe1.plot(x,result1[i,:],label=label_list[i],color=color_list[i],linestyle=linestyle_list[i])
    axe2.plot(x, result2[i,:], label=label_list[i], color=color_list[i],linestyle=linestyle_list[i])

# axe1.set_ylim([0.5,0.9])
# axe2.set_ylim([0.2,0.5])
axe1.legend(loc='lower right',fontsize=fontsize*2)
axe2.legend(loc='lower right',fontsize=fontsize*2)
axe1.set_ylim([0.2,0.9])
axe2.set_ylim([0.1,0.45])


fig.subplots_adjust(left=0.06,bottom=0.06,right=0.98,top=0.98,wspace=0.3,hspace=0.4)
fig.show()
# filename = '../fig/E1 Cochlea/ E1 Channels,bandwidth factor and classification accuracy.tif'
filename = '../fig/E2/E2 Channels, SD range and classification accuracy.tif'
fig.savefig(filename,dpi=450)


# a= 10
# row_list = []
# for i in range(2,rows+1):
#     col_list=[]
#     for j in cols:
#         col_list.append(ws.cell(i, j).value)
#     row_list.append(col_list)
# result = np.array(row_list)

# fig = plt.figure(figsize=(9,2))
# fontsize=8
# title_lis=['(a)','(b)','(c)','(d)','(e)','(f)']
#
# for i in range(1,7):
#     axe = fig.add_subplot(1, 6, i)
#     x=[0.2,0.1,0.05]
#     y = result[(i-1)*3:i*3,0]
#     axe.plot(x,y)
#     axe.set_title(title_lis[i-1], color='black',y=-0.55)
#     axe.set_xlim([0,0.25])
#     axe.set_ylim([0, 1])
#     axe.set_xlabel('Bandwidth factor', fontsize=fontsize)
#     axe.set_ylabel('Train accuracy', fontsize=fontsize)
#     # axe.set_ylim([0,100])
# # axe.set_ylabel('Time step', fontsize=fontsize)
# # axe.set_xlabel('Neuron index', fontsize=fontsize)
# fig.subplots_adjust(left=0.08,bottom=0.35,right=0.97,top=0.97,wspace=0.8,hspace=0.32)
# fig.show()
# filename = '../fig/E1/E1 Train accuracy and bandwidth factor.tif'
# fig.savefig(filename)
#
#
# a = 10
# sheet = wb['Sheet']
# sheet.title = 'result'
# # data_all = np.array([train_AA_par,test_AA_par]).transpose()
# for i in range(1, data_all.shape[0] + 1):
#     for j in range(1, data_all.shape[1] + 1):
#         wb['result'].cell(i, j, f'{data_all[i-1, j-1]:.4f}')
# wb.save('../tab/temp.xlsx')
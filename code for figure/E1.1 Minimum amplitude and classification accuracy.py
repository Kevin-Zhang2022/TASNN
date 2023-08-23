import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook


# wb = Workbook()
filename = 'E1.1 AN IHC/Amplitude range of AN'
wb = load_workbook(f'../tab/{filename}.xlsx')
ws = wb['Sheet1']

# rows = ws.max_row
rows = [np.arange(2,12)]
cols = [2]

result1=[]
result2=[]
for row_group in rows:
    for ind_col in cols:
        temp1 = []
        temp2 = []
        for ind_row in row_group:
            temp1.append(ws.cell(ind_row, ind_col).value)
            temp2.append(ws.cell(ind_row, ind_col+1).value)
        result1.append(temp1)
        result2.append(temp2)
result1=np.array(result1)
result2=np.array(result2)

fig = plt.figure(figsize=(8,6))

fig.show()
fontsize=8
x=[0.01,0.02,0.03,0.05,0.08,0.1,0.2,0.3,0.5,0.8]
# title_lis=['(a)','(b)','(c)','(d)','(e)','(f)']
color_list=['r','k','gold','g','royalblue','deeppink','purple']
label_list=['Average CC before IC','Average CC after IC']
linestyle_list = ['solid','dotted','dashed','dashdot',(0, (1, 1)),(0, (5, 10)),(0, (3, 10, 1, 10, 1, 10))]

axe1 = fig.add_subplot(1, 1, 1)
# axe2 = fig.add_subplot(1, 2, 2)
# for i in range(0,2):
axe1.plot(x,result1[0,:],color=color_list[0],linestyle=linestyle_list[0],label='Train accuracy')
axe1.plot(x,result2[0,:],color=color_list[1],linestyle=linestyle_list[1],label='Test accuracy')

# axe2.plot(x,result1[2,:],color=color_list[0],linestyle=linestyle_list[0],label='Before IC')
# axe2.plot(x,result2[2,:],color=color_list[1],linestyle=linestyle_list[1],label='Before IC')
# axe2.plot(x, result2[i,:], label=label_list[i], color=color_list[i],linestyle=linestyle_list[i*2+1])

# axe1.set_ylim([0.5,0.9])
# axe2.set_ylim([0.2,0.5])
axe1.legend(loc='upper right',fontsize=fontsize*2)
# axe2.legend(loc='upper right',fontsize=fontsize)

fig.subplots_adjust(left=0.05,bottom=0.06,right=0.98,top=0.98,wspace=0.3,hspace=0.4)
fig.show()
# filename = '../fig/E1 Cochlea/ E1 Channels,bandwidth factor and classification accuracy.tif'
filename = '../fig/E1.1 AN IHC/E1.1 Minimum amplitude and classification accuracy.tif'
fig.savefig(filename)


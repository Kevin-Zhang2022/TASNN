import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook


# wb = Workbook()
filename = 'bushy'
wb = load_workbook(f'../tab/E2/{filename}.xlsx')
ws = wb['Sheet2']

rows = ws.max_row
cols = [1,4]


row_list = []
for i in range(2,rows+1):
    col_list=[]
    for j in cols:
        col_list.append(ws.cell(i, j).value)
    row_list.append(col_list)
result = np.array(row_list)

fig = plt.figure(figsize=(10,2))
fontsize=8
title_lis=['(a)','(b)','(c)','(d)']

for i in range(1,4+1):
    axe = fig.add_subplot(1, 5, i)
    x=[100,250,500,1000]
    y = result[(i-1)*4:i*4,0]
    axe.plot(x,y)
    axe.set_title(title_lis[i-1], color='black',y=-0.5)
    axe.set_xlim([0,1100])
    axe.set_ylim([0.5, 1])
    axe.set_xlabel('Neuron number', fontsize=fontsize)
    axe.set_ylabel('Train accuracy', fontsize=fontsize)

x = [0.3281,0.9499,1.8511,2.8827]
y= [0.7959,0.7401,0.6970,0.6692]
axe = fig.add_subplot(1, 5, 5)
axe.plot(x, y)
axe.set_title('(e)', color='black', y=-0.5)
axe.set_xlim([0.2, 3])
axe.set_ylim([0.5, 1])
axe.set_xlabel('Average overlapping factor', fontsize=fontsize)
axe.set_ylabel('Average train accuracy', fontsize=fontsize)


fig.subplots_adjust(left=0.06,bottom=0.35,right=0.98,top=0.97,wspace=0.4,hspace=0.32)
fig.show()
filename = '../fig/E2/bushy overlapping factor trend.pdf'
fig.savefig(filename)


a = 10
# sheet = wb['Sheet']
# sheet.title = 'result'
# # data_all = np.array([train_AA_par,test_AA_par]).transpose()
# for i in range(1, data_all.shape[0] + 1):
#     for j in range(1, data_all.shape[1] + 1):
#         wb['result'].cell(i, j, f'{data_all[i-1, j-1]:.4f}')
# wb.save('../tab/temp.xlsx')
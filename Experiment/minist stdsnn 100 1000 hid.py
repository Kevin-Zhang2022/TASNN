import numpy as np
from cla.Data_set import Data_set as ds
from torch.utils.data import DataLoader
from cla.Net import Net
import torch
import torch.nn as nn
from cla.Show import Show as show
from Fun.norm import norm
from cla.Data_process import Data_process as dp
import matplotlib.pyplot as plt
# imports
import snntorch as snn
from snntorch import spikeplot as splt
from snntorch import spikegen

from torch.utils.data import DataLoader
from torchvision import datasets, transforms
import itertools

##
batch_size = 128
data_path='/data/mnist'

dtype = torch.float
device = torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")

# Define a transform
transform = transforms.Compose([
            transforms.Resize((28, 28)),
            transforms.Grayscale(),
            transforms.ToTensor(),
            transforms.Normalize((0,), (1,))])

mnist_train = datasets.MNIST(data_path, train=True, download=True, transform=transform)
mnist_test = datasets.MNIST(data_path, train=False, download=True, transform=transform)

train_loader = DataLoader(mnist_train, batch_size=batch_size, shuffle=True, drop_last=True)
test_loader = DataLoader(mnist_test, batch_size=batch_size, shuffle=True, drop_last=True)
##
# Network Architecture
num_inputs = 28*28
num_hidden = 1000
num_outputs = 10

# Temporal Dynamics
num_steps = 25
beta = 0.95
# Define Network
class Net(nn.Module):
    def __init__(self):
        super().__init__()

        # Initialize layers
        self.fc1 = nn.Linear(num_inputs, num_hidden)
        self.lif1 = snn.Leaky(beta=beta)
        self.fc2 = nn.Linear(num_hidden, num_outputs)
        self.lif2 = snn.Leaky(beta=beta)

    def forward(self, x):
        # Initialize hidden states at t=0
        mem1 = self.lif1.init_leaky()
        mem2 = self.lif2.init_leaky()

        # Record the final layer
        spk2_rec = []
        mem2_rec = []

        for step in range(num_steps):
            cur1 = self.fc1(x)
            spk1, mem1 = self.lif1(cur1, mem1)
            cur2 = self.fc2(spk1)
            spk2, mem2 = self.lif2(cur2, mem2)
            spk2_rec.append(spk2)
            mem2_rec.append(mem2)

        return torch.stack(spk2_rec, dim=0), torch.stack(mem2_rec, dim=0)
# Load the network onto CUDA if available
net = Net().to(device)
##
def print_batch_accuracy(data, targets, train=False):
    output, _ = net(data.view(batch_size, -1))
    _, idx = output.sum(dim=0).max(1)
    acc = np.mean((targets == idx).detach().cpu().numpy())

    if train:
        print(f"Train set accuracy for a single minibatch: {acc*100:.2f}%")
    else:
        print(f"Test set accuracy for a single minibatch: {acc*100:.2f}%")

def train_printer():
    print(f"Epoch {epoch}, Iteration {iter_counter}")
    print(f"Train Set Loss: {loss_hist[counter]:.2f}")
    print(f"Test Set Loss: {test_loss_hist[counter]:.2f}")
    print_batch_accuracy(data, targets, train=True)
    print_batch_accuracy(test_data, test_targets, train=False)
    print("\n")

loss = nn.CrossEntropyLoss()
optimizer = torch.optim.Adam(net.parameters(), lr=5e-4, betas=(0.9, 0.999))


num_epochs = 1
loss_hist = []
test_loss_hist = []
counter = 0

# Outer training loop
for epoch in range(num_epochs):
    iter_counter = 0
    train_batch = iter(train_loader)

    # Minibatch training loop
    for data, targets in train_batch:
        data = data.to(device)
        targets = targets.to(device)

        # plt.imshow(data[0,0,:,:])
        # forward pass
        net.train()
        spk_rec, mem_rec = net(data.view(batch_size, -1))
        plt.plot(mem_rec[:,0,:].detach().numpy())
        # initialize the loss & sum over time
        loss_val = torch.zeros((1), dtype=dtype, device=device)
        for step in range(num_steps):
            loss_val += loss(mem_rec[step], targets)

        # Gradient calculation + weight update
        optimizer.zero_grad()
        loss_val.backward()
        optimizer.step()

        # Store loss history for future plotting
        loss_hist.append(loss_val.item())

        # Test set
        with torch.no_grad():
            net.eval()
            test_data, test_targets = next(iter(test_loader))
            test_data = test_data.to(device)
            test_targets = test_targets.to(device)

            # Test set forward pass
            test_spk, test_mem = net(test_data.view(batch_size, -1))

            # Test set loss
            test_loss = torch.zeros((1), dtype=dtype, device=device)
            for step in range(num_steps):
                test_loss += loss(test_mem[step], test_targets)
            test_loss_hist.append(test_loss.item())

            # Print train/test loss/accuracy
            if counter % 50 == 0:
                train_printer()
            counter += 1
            iter_counter +=1


##

net = Net().to(device)

large_cycle = 10  # 10
reps = 5  # 5
epochs = 12  # 20
sc_step_size = 3  # 5
mode='ASNN'


window = 50
stride = 25
# train_loader = torch.load('train_loader.pth')
device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
data_type = torch.float32
loss = nn.CrossEntropyLoss()

# train
train_AA_i = []
train_AL_i = []
test_AA_i = []
test_AL_i = []
for i in range(large_cycle):
    dp.create_datalist(path2audio='../data/car_diagnostics/audio')
    train_dataset = ds(data_list_path='../data/car_diagnostics/train_list.csv',
                               max_duration=1, sample_rate=sample_rate, use_dB_normalization=False)
    test_dataset = ds(data_list_path='../data/car_diagnostics/test_list.csv',
                              max_duration=1, sample_rate=sample_rate, use_dB_normalization=False)
    train_loader = DataLoader(dataset=train_dataset, shuffle=True, batch_size=batch_size)
    test_loader = DataLoader(dataset=test_dataset, shuffle=True, batch_size=batch_size)

    net = Net(in_feature=channels, mode=mode)
    optimizer = torch.optim.Adam(net.parameters(), lr=0.1, betas=(0.9, 0.999))
    scheduler = torch.optim.lr_scheduler.StepLR(optimizer, step_size=sc_step_size, gamma=0.1)

    process='train'
    for epoch in range(epochs):
        for batch_id, (audio, label) in enumerate(train_loader):

            data = dp.scale(audio, mode='-11')
            data, cf = dp.gtfb(data, band_width=band_width, channels=channels, sam_rate=sample_rate)
            data = dp.ave_amp(data,window=window,stride=stride)

            out = net(data)

            loss_val = show.loss_val(loss,out,label,mode=mode)
            acc = show.acc(out, label, mode=mode)

            optimizer.zero_grad()
            loss_val.backward(retain_graph=True)
            optimizer.step()


            # if mode == 'std_snn':
            #     _,predicted = spk.sum(dim=1).max(1)
            #     # _, predicted = out.max(1)
            #     acc= np.mean((label == predicted).detach().numpy())
            #     # loss_val = loss_val.detach().numpy()[0]
            # else:
            #     _, predicted = out.max(1)
            #     acc = np.mean((label == predicted).detach().numpy())
            show.print_process(i,acc,epoch,loss_val,rep=0,process=process)
            # print(f"lr: {optimizer.state_dict()['param_groups'][0]['lr']:.2f}")
            # print('epoch:', epoch,'acc:%.2f',correct/total,'loss:',loss_val)
        scheduler.step()

    process='test train'
    train_loss = []
    train_acc = []
    for rep in range(reps):
        for batch_id, (audio, label) in enumerate(train_loader):
            data = dp.scale(audio, mode='-11')
            data, cf = dp.gtfb(data, band_width=band_width, channels=channels, sam_rate=sample_rate)
            data = dp.ave_amp(data,window=window,stride=stride)
            out = net(data)

            loss_val = show.loss_val(loss, out, label, mode=mode)
            acc = show.acc(out,label,mode=mode)

            train_acc.append(acc)
            train_loss.append(loss_val.item())
            show.print_process(i, acc, epoch, loss_val, rep=rep, process=process)
            # print(f"test train i: {i:d}", f"rep: {rep:d}", f"acc: {acc:.2f}", f"loss: {loss_val.item():.2f}")
    train_AA_i.append(np.mean(np.array(train_acc)))
    train_AL_i.append(np.mean(np.array(train_loss)))

    process='test test'
    test_loss = []
    test_acc = []
    for rep in range(reps):
        for batch_id, (audio, label) in enumerate(test_loader):
            data = dp.scale(audio, mode='-11')
            data, cf = dp.gtfb(data, band_width=band_width, channels=channels, sam_rate=sample_rate)
            data = dp.ave_amp(data,window=window,stride=stride)

            out = net(data)
            loss_val = show.loss_val(loss, out, label, mode=mode)
            acc = show.acc(out,label,mode=mode)
            show.print_process(i, acc, epoch, loss_val, rep=rep, process=process)

            test_acc.append(acc)
            test_loss.append(loss_val.detach().numpy())
    test_AA_i.append(np.mean(np.array(test_acc)))
    test_AL_i.append(np.mean(np.array(test_loss)))
a=10

plt.plot(out[1][0,:,:].detach().numpy())

from openpyxl import Workbook

wb = Workbook()
# wb = load_workbook('fuck.xlsx')
# print(wb.sheetnames)
sheet = wb['Sheet']
sheet.title = 'result'
# wb.create_sheet('test train loss')
# wb.create_sheet('test test acc')
# wb.create_sheet('test test loss')
# sheet_list = wb.sheetnames

data_all = np.array([train_AA_i,train_AL_i,test_AA_i,test_AL_i])
# a = np.random.random([9, 9])

for i in range(1, data_all.shape[0] + 1):
    for j in range(1, data_all.shape[1] + 1):
        wb['result'].cell(i, j, f'{data_all[i-1, j-1]:.4f}')
wb.save('../tab/std_snn 2layer 1000hid.xlsx')

a=10

acc_all = np.array(acc_all)
test_acc = np.mean(acc_all)
loss_all = np.array(loss_all)
test_loss = np.mean(loss_all)
print(f'i:{i:d}',f'test_acc:{test_acc:.2f}',f'test_loss:{test_loss:.2f}')


a =torch.rand(2,3,2,3)
layer = torch.nn.LayerNorm([2,3],elementwise_affine=False)
b = layer(a)





##


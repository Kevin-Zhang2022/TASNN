import numpy as np
from cla.Data_set import Data_set as ds
from torch.utils.data import DataLoader
from cla.Net import Net
import torch
import torch.nn as nn
from cla.Show import Show as show
from openpyxl import Workbook
from cla.Data_process import Data_process as dp
import matplotlib.pyplot as plt


# ran = [0.45,0.65]
# a= ran[0]+np.random.random([8,5])*(ran[1]-ran[0])
#
# wb = Workbook()
# # wb = load_workbook('fuck.xlsx')
# # print(wb.sheetnames)
# sheet = wb['Sheet']
# sheet.title = 'result'
#
# data_all = a
# # a = np.random.random([9, 9])
#
# for i in range(1, data_all.shape[0] + 1):
#     for j in range(1, data_all.shape[1] + 1):
#         wb['result'].cell(i, j, f'{data_all[i-1, j-1]:.4f}')
# wb.save('../tab/temp.xlsx')

sample_rate = 10000
band_width = [0, int(0.5*sample_rate)]

channels = 200
hidden = 10
out_feature=10

batch_size = 20
large_cycle = 10  # 10
reps = 3  # 5
epochs = 12  # 20
sc_step_size = 3  # 5
mode='SSNN'  # std_snn for pre-spike snn/ ASNN for post-spike snn


window = 50
stride = int(0.5*window)
# train_loader = torch.load('train_loader.pth')
device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
data_type = torch.float32
loss = nn.CrossEntropyLoss()

# train
train_AA_i = []
train_AL_i = []
test_AA_i = []
test_AL_i = []
for i in range(large_cycle):
    dp.create_datalist(path2audio='../data/car_diagnostics/audio')
    train_dataset = ds(data_list_path='../data/car_diagnostics/train_list.csv',
                               max_duration=1, sample_rate=sample_rate, use_dB_normalization=False)
    test_dataset = ds(data_list_path='../data/car_diagnostics/test_list.csv',
                              max_duration=1, sample_rate=sample_rate, use_dB_normalization=False)
    train_loader = DataLoader(dataset=train_dataset, shuffle=True, batch_size=batch_size)
    test_loader = DataLoader(dataset=test_dataset, shuffle=True, batch_size=batch_size)

    net = Net(in_feature=channels, hidden=hidden, out_feature=out_feature, mode=mode)
    optimizer = torch.optim.Adam(net.parameters(), lr=0.1, betas=(0.9, 0.999))
    scheduler = torch.optim.lr_scheduler.StepLR(optimizer, step_size=sc_step_size, gamma=0.1)

    process='train'
    for epoch in range(epochs):
        for batch_id, (audio, label) in enumerate(train_loader):

            data = dp.scale(audio, mode='-11')
            data, cf = dp.gtfb(data, band_width=band_width, channels=channels, sam_rate=sample_rate)
            data = dp.ave_amp(data,window=window,stride=stride)

            out = net(data)

            loss_val = show.loss_val(loss,out,label,mode=mode)
            acc = show.acc(out, label, mode=mode)

            optimizer.zero_grad()
            loss_val.backward(retain_graph=True)
            optimizer.step()


            # if mode == 'std_snn':
            #     _,predicted = spk.sum(dim=1).max(1)
            #     # _, predicted = out.max(1)
            #     acc= np.mean((label == predicted).detach().numpy())
            #     # loss_val = loss_val.detach().numpy()[0]
            # else:
            #     _, predicted = out.max(1)
            #     acc = np.mean((label == predicted).detach().numpy())
            show.print_process(i,acc,epoch,loss_val,rep=0,process=process)
            # print(f"lr: {optimizer.state_dict()['param_groups'][0]['lr']:.2f}")
            # print('epoch:', epoch,'acc:%.2f',correct/total,'loss:',loss_val)
        scheduler.step()

    process='test train'
    train_loss = []
    train_acc = []
    for rep in range(reps):
        for batch_id, (audio, label) in enumerate(train_loader):
            data = dp.scale(audio, mode='-11')
            data, cf = dp.gtfb(data, band_width=band_width, channels=channels, sam_rate=sample_rate)
            data = dp.ave_amp(data,window=window,stride=stride)
            out = net(data)

            loss_val = show.loss_val(loss, out, label, mode=mode)
            acc = show.acc(out,label,mode=mode)

            train_acc.append(acc)
            train_loss.append(loss_val.item())
            show.print_process(i, acc, epoch, loss_val, rep=rep, process=process)
            # print(f"test train i: {i:d}", f"rep: {rep:d}", f"acc: {acc:.2f}", f"loss: {loss_val.item():.2f}")
    train_AA_i.append(np.mean(np.array(train_acc)))
    train_AL_i.append(np.mean(np.array(train_loss)))

    process='test test'
    test_loss = []
    test_acc = []
    for rep in range(reps):
        for batch_id, (audio, label) in enumerate(test_loader):
            data = dp.scale(audio, mode='-11')
            data, cf = dp.gtfb(data, band_width=band_width, channels=channels, sam_rate=sample_rate)
            data = dp.ave_amp(data,window=window,stride=stride)

            out = net(data)
            loss_val = show.loss_val(loss, out, label, mode=mode)
            acc = show.acc(out,label,mode=mode)
            show.print_process(i, acc, epoch, loss_val, rep=rep, process=process)

            test_acc.append(acc)
            test_loss.append(loss_val.detach().numpy())
    test_AA_i.append(np.mean(np.array(test_acc)))
    test_AL_i.append(np.mean(np.array(test_loss)))
a=10

plt.plot(out[1][0,:,:].detach().numpy())



wb = Workbook()
# wb = load_workbook('fuck.xlsx')
# print(wb.sheetnames)
sheet = wb['Sheet']
sheet.title = 'result'
# wb.create_sheet('test train loss')
# wb.create_sheet('test test acc')
# wb.create_sheet('test test loss')
# sheet_list = wb.sheetnames

data_all = np.array([train_AA_i,train_AL_i,test_AA_i,test_AL_i])
# a = np.random.random([9, 9])

for i in range(1, data_all.shape[0] + 1):
    for j in range(1, data_all.shape[1] + 1):
        wb['result'].cell(i, j, f'{data_all[i-1, j-1]:.4f}')
wb.save('../tab/temp.xlsx')

a=10

# acc_all = np.array(acc_all)
# test_acc = np.mean(acc_all)
# loss_all = np.array(loss_all)
# test_loss = np.mean(loss_all)
# print(f'i:{i:d}',f'test_acc:{test_acc:.2f}',f'test_loss:{test_loss:.2f}')






##


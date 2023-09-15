import numpy as np
from cla.Data_set import Data_set as ds
from torch.utils.data import DataLoader
from cla.Net import Net
import torch
import torch.nn as nn
from cla.Show import Show as show
from Fun.norm import norm
from cla.Data_process import Data_process as dp
import matplotlib.pyplot as plt
from openpyxl import Workbook
from cla.GlobalParameter import GlobalParameter as gp
import random
from scipy import stats
import os

# def seed_torch(seed=10):
#     np.random.seed(seed)
#     os.environ['PYTHONHASHSEED'] = str(seed)
#     np.random.seed(seed)
#     torch.manual_seed(seed)
#     # torch.backends.cudnn.benchmark = False
#     # torch.backends.cudnn.deterministic = True
# seed_torch()

# def setup_seed(seed):
#     torch.manual_seed(seed)
#     torch.cuda.manual_seed_all(seed)
#     np.random.seed(seed)
#     random.seed(seed)
#     torch.backends.cudnn.deterministic = True
#
#
# # 设置随机数种子
# setup_seed(20)


large_cycle = 10  # 10
reps = 3  # 5
epochs = 12  # 20
sc_step_size = 3  # 5


window = gp.window
stride = gp.stride
# train_loader = torch.load('train_loader.pth')
device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
data_type = torch.float32
loss = nn.CrossEntropyLoss()


batch_size = 40

channels = gp.channels
hidden = gp.hidden

task = gp.task
out_features = gp.out_features

mode='mh' # 'lstm' 'gru' 'rnn' 'fc' 'mh'

train_AA_par = []
test_AA_par =[]
# train

train_AA_i = []
train_AL_i = []
test_AA_i = []
test_AL_i = []
for i in range(large_cycle):
    dp.create_datalist(path2audio='../data/'+task+'/audio')
    train_dataset = ds(data_list_path='../data/'+task+'/train_list.csv',
                       max_duration=1, sample_rate=gp.sample_rate, use_dB_normalization=False)
    test_dataset = ds(data_list_path='../data/'+task+'/test_list.csv',
                      max_duration=1, sample_rate=gp.sample_rate, use_dB_normalization=False)
    train_loader = DataLoader(dataset=train_dataset, shuffle=True, batch_size=batch_size)
    test_loader = DataLoader(dataset=test_dataset, shuffle=True, batch_size=20)
    net = Net(in_features=channels, hidden=hidden, out_features=out_features, mode=mode)
    optimizer = torch.optim.Adam(net.parameters(), lr=0.1, betas=(0.9, 0.999))
    scheduler = torch.optim.lr_scheduler.StepLR(optimizer, step_size=sc_step_size, gamma=0.1)

    process='train'
    for epoch in range(epochs):
        for batch_id, (audio, label) in enumerate(train_loader):
            data = dp.scale(audio, mode='-11')
            # data, cf = dp.gtfb(data, band_width=gp.band_width, channels=channels, sam_rate=gp.sample_rate)
            # data = dp.ave_amp(data,window=window,stride=stride)

            # plt.plot(audio[0,:])
            # audio[0,500]
            # data[0,500]
            # data[0,200,25]

            # max,_ = torch.max(data,dim=1)
            # plt.plot(max[6,:].detach().numpy())

            out = net(data)
            # plt.imshow(out[0][0,:,:].detach().numpy(),aspect='auto')
            # IF = torch.sum(torch.abs(net.fc2.weight),dim=0)
            # plt.plot(IF.detach().numpy())
            # loss_val = torch.zeros(1)
            # for t in range(mem.shape[1]):
            #     loss_val += loss(mem[:, t, :], label)
            loss_val = show.loss_val(loss,out,label,mode=mode)
            acc = show.acc(out, label, mode=mode)

            # a = net.fc0.weight.grad
            optimizer.zero_grad()
            loss_val.backward(retain_graph=True)
            optimizer.step()
            # a = net.fc2.weight.grad
            # a = net.fc1.e_linear1.w.grad
            # a = net.rs1.plinear1.w.grad
            # plt.imshow(a.detach().numpy(),aspect='auto')

            # plt.figure()
            # b = net.rs1.plinear1.w[0,:].detach().numpy()
            # plt.ylim([-0.8,0.8])
            # plt.plot(b)
            # plt.title('rs1 w 1000 1 LAYER')
            # ind = torch.where(a>0)

            show.print_process(i,acc,epoch,loss_val,rep=0,process=process)
            # print(f"lr: {optimizer.state_dict()['param_groups'][0]['lr']:.2f}")
            # print('epoch:', epoch,'acc:%.2f',correct/total,'loss:',loss_val)
        scheduler.step()

    process='test train'
    train_loss = []
    train_acc = []
    for rep in range(reps):
        for batch_id, (audio, label) in enumerate(train_loader):
            data = dp.scale(audio, mode='-11')
            # data, cf = dp.gtfb(data, band_width=band_width, channels=channels, sam_rate=sample_rate)
            # data = dp.ave_amp(data,window=window,stride=stride)
            out = net(data)

            loss_val = show.loss_val(loss, out, label, mode=mode)
            acc = show.acc(out,label,mode=mode)

            train_acc.append(acc)
            train_loss.append(loss_val.item())
            show.print_process(i, acc, epoch, loss_val, rep=rep, process=process)
            # print(f"test train i: {i:d}", f"rep: {rep:d}", f"acc: {acc:.2f}", f"loss: {loss_val.item():.2f}")
    train_AA_i.append(np.mean(np.array(train_acc)))
    train_AL_i.append(np.mean(np.array(train_loss)))

    process='test test'
    test_loss = []
    test_acc = []
    for rep in range(reps):
        for batch_id, (audio, label) in enumerate(test_loader):
            data = dp.scale(audio, mode='-11')
            # data, cf = dp.gtfb(data, band_width=band_width, channels=channels, sam_rate=sample_rate)
            # data = dp.ave_amp(data,window=window,stride=stride)

            out = net(data)
            loss_val = show.loss_val(loss, out, label, mode=mode)
            acc = show.acc(out,label,mode=mode)
            show.print_process(i, acc, epoch, loss_val, rep=rep, process=process)

            test_acc.append(acc)
            test_loss.append(loss_val.detach().numpy())
    test_AA_i.append(np.mean(np.array(test_acc)))
    test_AL_i.append(np.mean(np.array(test_loss)))
train_AA_par.append(np.mean(train_AA_i))
test_AA_par.append(np.mean(test_AA_i))
a=10
# plt.plot(out[1][0,:,:].detach().numpy())
wb = Workbook()
# wb = load_workbook('fuck.xlsx')
# print(wb.sheetnames)
sheet = wb['Sheet']
sheet.title = 'result'
data_all = np.array([train_AA_par,test_AA_par]).transpose()
for i in range(1, data_all.shape[0] + 1):
    for j in range(1, data_all.shape[1] + 1):
        wb['result'].cell(i, j, f'{data_all[i-1, j-1]:.4f}')
wb.save('../tab/temp.xlsx')

# a=10
# a = net.bushy1.elinear_narrow.w
# b = net.bushy1.elinear_medium.w
# c = net.bushy1.elinear_broad.w
# d=1+a+b+c

# plt.figure()
# plt.plot(d[0,:].detach().numpy())
# plt.plot(net.bushy1.elinear_narrow.w[0,:].detach().numpy())
# plt.figure()
# plt.plot(net.bushy1.elinear_medium.w[0,:].detach().numpy())
# plt.figure()
# plt.plot(net.bushy1.elinear_broad.w[0,:].detach().numpy())
# plt.show()


# torch.save(net,'net.pkl')
    # acc_all = np.array(acc_all)
    # test_acc = np.mean(acc_all)
    # loss_all = np.array(loss_all)
    # test_loss = np.mean(loss_all)
    # print(f'i:{i:d}',f'test_acc:{test_acc:.2f}',f'test_loss:{test_loss:.2f}')


# a =torch.rand(2,3,2,3)
# layer = torch.nn.LayerNorm([2,3],elementwise_affine=False)
# b = layer(a)



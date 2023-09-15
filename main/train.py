import numpy as np
from cla.Data_set import Data_set as ds
from torch.utils.data import DataLoader
from cla.Net import Net
import torch
import torch.nn as nn
from cla.Show import Show as show
from Fun.norm import norm
from cla.Data_process import Data_process as dp
import matplotlib.pyplot as plt
from openpyxl import Workbook
from cla.GlobalParameter import GlobalParameter as gp
import random
from scipy import stats
import os

# def seed_torch(seed=10):
#     np.random.seed(seed)
#     os.environ['PYTHONHASHSEED'] = str(seed)
#     np.random.seed(seed)
#     torch.manual_seed(seed)
#     # torch.backends.cudnn.benchmark = False
#     # torch.backends.cudnn.deterministic = True
# seed_torch()

# def setup_seed(seed):
#     torch.manual_seed(seed)
#     torch.cuda.manual_seed_all(seed)
#     np.random.seed(seed)
#     random.seed(seed)
#     torch.backends.cudnn.deterministic = True
#
#
# # 设置随机数种子
# setup_seed(20)


# check hyper parameters
# bandwidth_factor   hidden batch size channel
#
#
large_cycle = 5  # 5 10
reps = 3  # 5 3
epochs = 12  # 20 12
sc_step_size = 3  # 5 3

# large_cycle = 1  # 10
# reps = 1  # 5 3
# epochs = 1  # 20 12
# sc_step_size = 1  # 5 3

window_list=[30,50,100,200,300,500,800,1000]
stride_list = [0.1,0.2,0.3,0.5,0.8,1]

sample_rate = 10000
frequency_range=[5,5000]
# train_loader = torch.load('train_loader.pth')
device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
data_type = torch.float32
loss = nn.CrossEntropyLoss()
train_batchsize_dic = {'car_diagnostics': 40,'esc_50': 100, 'urban_sound8k': 100}
test_batchsize_dic = {'car_diagnostics': 20,'esc_50': 100, 'urban_sound8k': 100}

input_list = [20,50,100,150,200,300,500]

hidden_list=[100,200,300,500,800,1000]

model_list=['lstm','gru','rnn','mh']
task_list = ['car_diagnostics', 'esc_50', 'urban_sound8k','music_genre']
out_dic = {'car_diagnostics': 10,'esc_50': 50, 'urban_sound8k': 10,'music_genre': 9}
bandwidth_factor_list=[0.01,0.02,0.03,0.05,0.07,0.1,0.15,0.2,0.5,1]

ans_list=[2,3,5,7,8,10,12,15,17,20]
an_range_list=[0.01,0.02,0.03,0.05,0.07,0.09,0.1,0.2,0.3,0.5]

bushy_range_list = [0.3,0.5,0.8,5,8,10,20]

ic_mode_list=[True, False]

one_list0=[1]
one_list1=[1]


for ind_0 in range(len(one_list0)):
    train_AA_par = []
    test_AA_par = []
    bushy_cc_train_par = []
    ic_cc_train_par = []
    bushy_cc_test_par = []
    ic_cc_test_par = []
    for ind_1 in range(len(one_list1)):
        window = 50  # window_list=[30,50,100,200,300,500,800,1000]
        stride = int(0.5*window)  # [0.1,0.2,0.3,0.5,0.8,1]

        model = 'mh' # default 'mh' model_list=['lstm','gru','rnn','mh']
        task= 'car_diagnostics'  # task_list = ['car_diagnostics', 'esc_50', 'urban_sound8k','music_genre']
        out_features = out_dic[task]  # 10 10 50
        train_batchsize = train_batchsize_dic[task]
        test_batchsize = test_batchsize_dic[task]

        in_features = 200   # input_list = [20,50,100,150,200,300,500] default 200
        hidden= 500  # hidden_list=[100,200,300,500,800,1000] default 500

        ans=10  # ans_list=[2,3,5,7,8,10,12,15,17,20]
        an_range = [0.03,1]  # an_range_list = [0.01,0.02,0.03,0.05,0.07,0.1,0.2,0.3,0.5,0.8,0.99] default 0.03

        bushy_range=[0.2,2]  # bushy_range_list = [0.3,0.5,0.8,1,1.2,1.5,1.7,2,2.2,2.5,2.8,3] default 2
        bandwidth_factor = 0.05  # bandwidth_factor_list=[0.01,0.02,0.03, 0.05, 0.07, 0.1, 0.15,0.2,0.5,1] default 0.05

        ic_mode=True  # ic_mode_list=[True, False] default True

        train_AA_i = []
        train_AL_i = []
        test_AA_i = []
        test_AL_i = []

        train_ACC_bushy_i=[]
        train_ACC_ic_i=[]
        test_ACC_bushy_i = []
        test_ACC_ic_i = []

        for ind_largcycle in range(large_cycle):
            dp.create_datalist(path2audio='../data/'+task+'/audio')
            train_dataset = ds(data_list_path='../data/'+task+'/train_list.csv',
                               max_duration=1, sample_rate=sample_rate, use_dB_normalization=False)
            test_dataset = ds(data_list_path='../data/'+task+'/test_list.csv',
                              max_duration=1, sample_rate=sample_rate, use_dB_normalization=False)
            train_loader = DataLoader(dataset=train_dataset, shuffle=True, batch_size=train_batchsize)
            test_loader = DataLoader(dataset=test_dataset, shuffle=True, batch_size=test_batchsize)
            net = Net(in_features=in_features, hidden=hidden, out_features=out_features, model=model,
                      bandwidth_factor=bandwidth_factor, frequency_range=frequency_range, sample_rate=sample_rate,
                      window=window,stride=stride,
                      ans=ans, an_range=an_range,
                      bushy_range=bushy_range,ic_mode=ic_mode)
            optimizer = torch.optim.Adam(net.parameters(), lr=0.1, betas=(0.9, 0.999))
            scheduler = torch.optim.lr_scheduler.StepLR(optimizer, step_size=sc_step_size, gamma=0.1)

            process='train'
            for epoch in range(epochs):
                for batch_id, (audio, label) in enumerate(train_loader):
                    data = dp.normalize(audio)
                    out = net(data)
                    # plt.imshow(out[0][0,:,:].detach().numpy(),aspect='auto')
                    # IF = torch.sum(torch.abs(net.fc2.weight),dim=0)
                    # plt.plot(IF.detach().numpy())
                    # loss_val = torch.zeros(1)
                    # for t in range(mem.shape[1]):
                    #     loss_val += loss(mem[:, t, :], label)
                    loss_val = show.loss_val(loss,out,label)
                    acc = show.acc(out, label)

                    # a = net.fc0.weight.grad
                    optimizer.zero_grad()
                    loss_val.backward(retain_graph=True)
                    optimizer.step()

                    show.print_process(ind_largcycle,acc,epoch,loss_val,rep=0,process=process)

                scheduler.step()

            # torch.save(net,'../net_for_figtab/average cc net.pkl')
            process='test train'
            train_loss = []
            train_acc = []
            bushy_cc_train = []
            ic_cc_train = []
            for rep in range(reps):
                for batch_id, (audio, label) in enumerate(train_loader):
                    data = dp.normalize(audio)
                    out = net(data)
                    # torch.save(net,'../net_for_figtab/bushy_ic bushy_range3.pkl')
                    loss_val = show.loss_val(loss, out, label)
                    acc = show.acc(out,label)

                    train_acc.append(acc)
                    train_loss.append(loss_val.item())
                    show.print_process(ind_largcycle, acc, epoch, loss_val, rep=rep, process=process)

                    bushy_cc_train.append(dp.average_ccmat(net.spk_bushy_stack.transpose(1,2)))
                    ic_cc_train.append(dp.average_ccmat(net.spk_ic_stack.transpose(1,2)))
            train_ACC_bushy_i.append(np.mean(np.array(bushy_cc_train)))
            train_ACC_ic_i.append(np.mean(np.array(ic_cc_train)))
            train_AA_i.append(np.mean(np.array(train_acc)))
            train_AL_i.append(np.mean(np.array(train_loss)))

            process='test test'
            test_loss = []
            test_acc = []
            bushy_cc_test = []
            ic_cc_test = []
            for rep in range(reps):
                for batch_id, (audio, label) in enumerate(test_loader):
                    data = dp.normalize(audio)
                    out = net(data)
                    loss_val = show.loss_val(loss, out, label)
                    acc = show.acc(out,label)
                    show.print_process(ind_largcycle, acc, epoch, loss_val, rep=rep, process=process)

                    test_acc.append(acc)
                    test_loss.append(loss_val.detach().numpy())
                    bushy_cc_test.append(dp.average_ccmat(net.spk_bushy_stack.transpose(1,2)))
                    ic_cc_test.append(dp.average_ccmat(net.spk_ic_stack.transpose(1,2)))
            test_ACC_bushy_i.append(np.mean(np.array(bushy_cc_test)))
            test_ACC_ic_i.append(np.mean(np.array(ic_cc_test)))
            test_AA_i.append(np.mean(np.array(test_acc)))
            test_AL_i.append(np.mean(np.array(test_loss)))
        train_AA_par.append(np.mean(train_AA_i))
        test_AA_par.append(np.mean(test_AA_i))

        bushy_cc_train_par.append(np.mean(train_ACC_bushy_i))
        ic_cc_train_par.append(np.mean(train_ACC_ic_i))
        bushy_cc_test_par.append(np.mean(test_ACC_bushy_i))
        ic_cc_test_par.append(np.mean(test_ACC_ic_i))


    wb = Workbook()
    # wb = load_workbook('fuck.xlsx')
    # print(wb.sheetnames)
    sheet = wb['Sheet']
    sheet.title = 'result'
    data_all = np.array([train_AA_par,test_AA_par,bushy_cc_train_par,ic_cc_train_par,bushy_cc_test_par,ic_cc_test_par]).transpose()
    # data_all = np.array([bushy_cc_train_par,ic_cc_train_par,bushy_cc_test_par,ic_cc_test_par]).transpose()

    for ind_row in range(1, data_all.shape[0] + 1):
        for ind_col in range(1, data_all.shape[1] + 1):
            wb['result'].cell(ind_row, ind_col, f'{data_all[ind_row-1, ind_col-1]:.4f}')
    wb.save('../tab/temp'+str(ind_0)+'.xlsx')


a=10
b=15
print('end')
# a=10
# a = net.bushy1.elinear_narrow.w
# b = net.bushy1.elinear_medium.w
# c = net.bushy1.elinear_broad.w
# d=1+a+b+c
# plt.figure()
# plt.plot(d[0,:].detach().numpy())
# plt.plot(net.bushy1.elinear_narrow.w[0,:].detach().numpy())
# plt.figure()
# plt.plot(net.bushy1.elinear_medium.w[0,:].detach().numpy())
# plt.figure()
# plt.plot(net.bushy1.elinear_broad.w[0,:].detach().numpy())
# plt.show()


# torch.save(net,'net.pkl')
    # acc_all = np.array(acc_all)
    # test_acc = np.mean(acc_all)
    # loss_all = np.array(loss_all)
    # test_loss = np.mean(loss_all)
    # print(f'i:{i:d}',f'test_acc:{test_acc:.2f}',f'test_loss:{test_loss:.2f}')


# a =torch.rand(2,3,2,3)
# layer = torch.nn.LayerNorm([2,3],elementwise_affine=False)
# b = layer(a)


